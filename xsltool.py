import openpyxl

from info import DeleteInfo, UpdateInfo, MergeInfo

DEFAULT_XLSX_FILE = 'info.xlsx'
# sheet 信息
DB_SHEET = '连接信息'
DELETE_ARCHIVE_SHEET = '待删除档案'
DELETE_TRACK_SHEET = '待删除轨迹'
UPDATE_ARCHIVE_SHEET = '待更新档案'
MERGE_ARCHIVE_SHEET = '待合并档案'


def read_xlsx(xls_file=DEFAULT_XLSX_FILE):
    return openpyxl.open(xls_file)


def read_sheet(workbook, sheet_name):
    return workbook[sheet_name]


# 读取数据库连接信息
def read_db_info(workbook):
    worksheet = read_sheet(workbook, DB_SHEET)
    if not worksheet:
        return
    ip = worksheet.cell(1, 2).value
    port = int(worksheet.cell(2, 2).value)
    user = worksheet.cell(3, 2).value
    password = worksheet.cell(4, 2).value
    db = worksheet.cell(5, 2).value
    return ip, port, user, password, db


# 读取待删除档案信息
def read_delete_archive(workbook):
    worksheet = read_sheet(workbook, DELETE_ARCHIVE_SHEET)
    if not worksheet:
        return
    delete_archives = []
    for c in worksheet.iter_rows(min_row=2, min_col=1, max_col=1):
        value = c[0].value
        if value:
            delete_archives.append(DeleteInfo(c[0].row, value))
    return delete_archives


# 读取待删除轨迹信息
def read_delete_tracks(workbook):
    worksheet = read_sheet(workbook, DELETE_TRACK_SHEET)
    if not worksheet:
        return
    tracks = []
    for c in worksheet.iter_rows(min_row=2, min_col=1, max_col=1):
        value = c[0].value
        if value:
            tracks.append(DeleteInfo(c[0].row, value))
    return tracks


# 读取待更新档案信息
def read_update_archive(workbook):
    worksheet = read_sheet(workbook, UPDATE_ARCHIVE_SHEET)
    if not worksheet:
        return
    archive_info = []
    for c in worksheet.iter_rows(min_row=2, min_col=1, max_col=9):
        values = c[0].value, c[1].value, c[2].value, c[3].value, c[4].value, c[5].value, c[6].value, c[7].value, c[
            8].value
        if not values[0] or not values[0].strip():
            continue
        archive_info.append(UpdateInfo(c[0].row, values))
    return archive_info


# 读取带合并档案的信息
def read_merge_archive(workbook):
    worksheet = read_sheet(workbook, MERGE_ARCHIVE_SHEET)
    if not worksheet:
        return
    merge_info = []
    for c in worksheet.iter_rows(min_row=2, min_col=1, max_col=2):
        from_archive = c[0].value
        to_archive = c[1].value
        if not from_archive and not to_archive:
            continue
        merge_info.append(MergeInfo(c[0].row, from_archive, to_archive))
    return merge_info


def write_delete_archive_result(workbook, delete_info, xls_file=DEFAULT_XLSX_FILE):
    worksheet = read_sheet(workbook, DELETE_ARCHIVE_SHEET)
    worksheet.cell(delete_info.result_row, delete_info.result_col).value = delete_info.result
    workbook.save(xls_file)


def write_delete_track_result(workbook, delete_info, xls_file=DEFAULT_XLSX_FILE):
    worksheet = read_sheet(workbook, DELETE_TRACK_SHEET)
    worksheet.cell(delete_info.result_row, delete_info.result_col).value = delete_info.result
    workbook.save(xls_file)


def write_update_archive_result(workbook, update_info, xls_file=DEFAULT_XLSX_FILE):
    worksheet = read_sheet(workbook, UPDATE_ARCHIVE_SHEET)
    worksheet.cell(update_info.result_row, update_info.result_col).value = update_info.result
    workbook.save(xls_file)


def write_merge_archive_result(workbook, merge_info, xls_file=DEFAULT_XLSX_FILE):
    worksheet = read_sheet(workbook, MERGE_ARCHIVE_SHEET)
    worksheet.cell(merge_info.result_row, merge_info.result_col).value = merge_info.result
    workbook.save(xls_file)
